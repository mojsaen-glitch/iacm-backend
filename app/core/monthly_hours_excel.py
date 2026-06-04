"""Crew Monthly Flight Hours — professional Excel workbook (openpyxl).

Five sheets: Monthly Matrix · Summary · Crew Without Hours · Blocked/Disconnected
· Raw Data. Styled with freeze panes, auto-filter, duty-type cell colours, column
widths, and HH:MM hour formatting. Built from the dict produced by
``monthly_hours.build_matrix`` — no DB access here.
"""
from __future__ import annotations
import calendar
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.core.monthly_hours import hm

_DH_LABEL = {"none": "Not counted (0%)", "half": "Half (50%)", "full": "Full (100%)"}


def _dh_label(v) -> str:
    return _DH_LABEL.get(v or "none", str(v))

# Palette
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_TITLE_FONT = Font(bold=True, size=15, color="1F4E78")
_SUB_FONT = Font(bold=True, size=11, color="404040")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_BOLD = Font(bold=True)
_OVER_FONT = Font(bold=True, color="9C0006")
_OPERATING_FILL = PatternFill("solid", fgColor="E2EFDA")   # green
_DEADHEAD_FILL = PatternFill("solid", fgColor="DDEBF7")    # blue
_STANDBY_FILL = PatternFill("solid", fgColor="E4DFEC")     # purple
_OVER_FILL = PatternFill("solid", fgColor="FFC7CE")        # red
_NOHOURS_FILL = PatternFill("solid", fgColor="F2F2F2")     # grey
_TOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")       # soft gold
_thin = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center")
_CENTER = Alignment(horizontal="center", vertical="center")


def _day_fill(legs: list[dict]):
    duties = {leg["duty_type"] for leg in legs}
    if "operating" in duties:
        return _OPERATING_FILL
    if "deadhead" in duties:
        return _DEADHEAD_FILL
    if "standby" in duties:
        return _STANDBY_FILL
    return None


def _route_chain(legs: list[dict]) -> str:
    parts: list[str] = []
    for leg in legs:
        r = (leg.get("route") or "").split("-")
        if len(r) == 2 and r[0]:
            if not parts:
                parts += [r[0], r[1]]
            else:
                if parts[-1] != r[0]:
                    parts.append(r[0])
                parts.append(r[1])
    return "-".join(parts)


def _day_text(day: dict) -> str:
    """Route chain + the day's (possibly overridden) credited hours. A pencil
    marks a manual override."""
    legs = day.get("legs", [])
    chain = _route_chain(legs)
    mark = "✎ " if day.get("override") else ""
    h = hm(day.get("day_hours"))
    body = chain if chain else ("MANUAL" if day.get("override") else "")
    return f"{mark}{body}" + (f"\n{h}" if h else "")


def build_workbook(matrix: dict, company_name: str = "") -> bytes:
    wb = Workbook()
    _sheet_matrix(wb.active, matrix, company_name)
    _sheet_summary(wb.create_sheet("Summary"), matrix, company_name)
    _sheet_calc_details(wb.create_sheet("Calculation Details"), matrix)
    _sheet_without(wb.create_sheet("Crew Without Hours"), matrix)
    _sheet_blocked(wb.create_sheet("Blocked - Disconnected"), matrix)
    _sheet_raw(wb.create_sheet("Raw Data"), matrix)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _included_of(leg: dict):
    """(Yes/No, short reason) for a matrix leg, mirroring the crediting policy."""
    duty = leg.get("duty_type")
    if duty == "operating":
        if (leg.get("block") or 0) <= 0:
            return "No", "Operating but duration missing — verify"
        return "Yes", "Operating — counted in full"
    return "No", {
        "deadhead": "Deadhead — not counted",
        "standby": "Standby — counts as a Standby day",
        "training": "Training — not counted",
        "observer": "Observer — not counted",
    }.get(duty, f"{duty} — not counted")


def _sheet_calc_details(ws, matrix: dict):
    """Every leg that fed each crew's hours — the audit trail behind the matrix."""
    headers = ["Crew", "Code", "Rank", "Date", "Flight No", "Duty", "Route", "Day Route",
               "Aircraft", "REG", "STD", "STA", "Duration", "Credited", "Included",
               "Reason", "Source", "Flight ID", "Assignment ID"]
    _header_row(ws, 1, headers)
    y, m = matrix["year"], matrix["month"]
    r = 2
    for row in matrix["rows"]:
        for day_str in sorted(row["days"], key=lambda x: int(x)):
            day = row["days"][day_str]
            chain = _route_chain(day["legs"])
            date = f"{y:04d}-{m:02d}-{int(day_str):02d}"
            for leg in day["legs"]:
                inc, reason = _included_of(leg)
                vals = [row["name"], row["code"], row["rank_code"] or row["rank"], date,
                        leg.get("flight_no", ""), leg.get("duty_type", ""), leg.get("route", ""),
                        chain, leg.get("aircraft_type", ""), leg.get("registration", ""),
                        leg.get("std", ""), leg.get("sta", ""), leg.get("block", 0),
                        leg.get("hours", 0), inc, reason, "flights.duration_hours",
                        leg.get("flight_id", ""), leg.get("assignment_id", "")]
                for ci, v in enumerate(vals, start=1):
                    ws.cell(r, ci, v)
                r += 1
    widths = [24, 12, 8, 12, 10, 10, 12, 16, 10, 12, 7, 7, 9, 9, 9, 30, 18, 20, 20]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(r - 1, 1)}"


def _sheet_matrix(ws, matrix: dict, company_name: str):
    ws.title = "Monthly Matrix"
    n = matrix["days_in_month"]
    rows = matrix["rows"]
    fixed = ["NAME", "CODE", "RANK", "A/C"]
    tail = ["1st Half", "2nd Half", "Month Total", "Flights", "Notes"]
    total_cols = len(fixed) + n + len(tail)

    ws.cell(1, 1, "Monthly Flight Hours Statistics").font = _TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(total_cols, 12))
    mname = calendar.month_name[matrix["month"]]
    dh = _dh_label(matrix.get("summary", {}).get("dh_credit"))
    ws.cell(2, 1, f"{company_name}   —   {mname} {matrix['year']}   ·   DH Crediting: {dh}").font = _SUB_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=min(total_cols, 12))

    hdr = 4
    headers = fixed + [str(d) for d in range(1, n + 1)] + tail
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(hdr, ci, h)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = _CENTER
        c.border = _BORDER

    r = hdr + 1
    for row in rows:
        no_hours = row["month_total"] <= 0
        name_c = ws.cell(r, 1, row["name"]); name_c.alignment = _LEFT
        ws.cell(r, 2, row["code"]).alignment = _CENTER
        ws.cell(r, 3, row["rank_code"] or row["rank"]).alignment = _CENTER
        ws.cell(r, 4, ", ".join(row["aircraft_types"])).alignment = _CENTER
        for d in range(1, n + 1):
            cell = ws.cell(r, 4 + d)
            day = row["days"].get(str(d))
            if day and (day["legs"] or day.get("day_hours")):
                cell.value = _day_text(day)
                cell.fill = _day_fill(day["legs"]) or _OPERATING_FILL
            cell.alignment = _WRAP
            cell.border = _BORDER
        base = 4 + n
        ws.cell(r, base + 1, hm(row["first_half"])).alignment = _CENTER
        ws.cell(r, base + 2, hm(row["second_half"])).alignment = _CENTER
        mt = ws.cell(r, base + 3, hm(row["month_total"]))
        mt.alignment = _CENTER
        mt.font = _BOLD
        mt.fill = _TOTAL_FILL
        if row["over_limit"]:
            mt.fill = _OVER_FILL
            mt.font = _OVER_FONT
        ws.cell(r, base + 4, row["flights_count"]).alignment = _CENTER
        notes = []
        if row["blocked"]:
            notes.append("GROUNDED")
        if row["deadhead_count"]:
            notes.append(f"DH:{row['deadhead_count']}")
        if row["standby_days"]:
            notes.append(f"STBY:{row['standby_days']}")
        ws.cell(r, base + 5, "  ".join(notes)).alignment = _LEFT
        if no_hours:
            for ci in range(1, 5):
                ws.cell(r, ci).fill = _NOHOURS_FILL
        r += 1

    # widths
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12
    for d in range(1, n + 1):
        ws.column_dimensions[get_column_letter(4 + d)].width = 13
    for i, _ in enumerate(tail):
        ws.column_dimensions[get_column_letter(4 + n + 1 + i)].width = 12
    ws.freeze_panes = "E5"
    last = r - 1 if r > hdr + 1 else hdr
    ws.auto_filter.ref = f"A{hdr}:{get_column_letter(total_cols)}{last}"
    ws.sheet_view.showGridLines = False


def _kv(ws, r, k, v, bold_val=False):
    ws.cell(r, 1, k).font = _BOLD
    c = ws.cell(r, 2, v)
    if bold_val:
        c.font = _BOLD
    return r + 1


def _sheet_summary(ws, matrix: dict, company_name: str):
    s = matrix["summary"]
    rows = matrix["rows"]
    ws.cell(1, 1, "Summary").font = _TITLE_FONT
    mname = calendar.month_name[matrix["month"]]
    ws.cell(2, 1, f"{company_name}   —   {mname} {matrix['year']}").font = _SUB_FONT

    r = 4
    for k, v in [
        ("Total Crew (in scope)", s["active_crew"] + s["crew_without_hours"]),
        ("Active Crew (with hours)", s["active_crew"]),
        ("Crew Without Hours", s["crew_without_hours"]),
        ("Blocked / Disconnected", s["blocked_crew"]),
        ("Total Monthly Hours", hm(s["total_hours"])),
        ("Highest Monthly Hours", hm(s["highest_hours"])),
        ("Lowest Monthly Hours", hm(s["lowest_hours"])),
        ("Total Flights", s["total_flights"]),
        ("Deadhead Hours", hm(s["deadhead_hours"])),
        ("Standby Days", s["standby_days"]),
        ("Compliance Warnings", s["compliance_warnings"]),
        ("DH Crediting Policy", _dh_label(s.get("dh_credit"))),
    ]:
        r = _kv(ws, r, k, v)

    with_hours = [x for x in rows if x["month_total"] > 0]
    top = sorted(with_hours, key=lambda x: -x["month_total"])[:10]
    bottom = sorted(with_hours, key=lambda x: x["month_total"])[:10]

    r += 1
    ws.cell(r, 1, "Top 10 by Hours").font = _SUB_FONT
    ws.cell(r, 4, "Bottom 10 by Hours").font = _SUB_FONT
    r += 1
    for i in range(max(len(top), len(bottom))):
        if i < len(top):
            ws.cell(r + i, 1, top[i]["name"])
            ws.cell(r + i, 2, hm(top[i]["month_total"]))
        if i < len(bottom):
            ws.cell(r + i, 4, bottom[i]["name"])
            ws.cell(r + i, 5, hm(bottom[i]["month_total"]))
    r += max(len(top), len(bottom)) + 1

    # by rank / aircraft / base
    by_rank, by_base = {}, {}
    by_ac = {}
    for x in rows:
        by_rank[x["rank_code"] or x["rank"]] = by_rank.get(x["rank_code"] or x["rank"], 0.0) + x["month_total"]
        by_base[x["base"]] = by_base.get(x["base"], 0.0) + x["month_total"]
        for d in x["days"].values():
            for leg in d["legs"]:
                if leg["duty_type"] == "operating" and leg["aircraft_type"]:
                    by_ac[leg["aircraft_type"]] = by_ac.get(leg["aircraft_type"], 0.0) + leg["hours"]

    for title, data in [("Hours by Rank", by_rank), ("Hours by Aircraft", by_ac), ("Hours by Base", by_base)]:
        r += 1
        ws.cell(r, 1, title).font = _SUB_FONT
        r += 1
        for k, v in sorted(data.items(), key=lambda kv: -kv[1]):
            if not k:
                continue
            ws.cell(r, 1, k)
            ws.cell(r, 2, hm(v))
            r += 1

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["D"].width = 26
    ws.column_dimensions["E"].width = 14


def _header_row(ws, r, headers):
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(r, ci, h)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = _CENTER
        c.border = _BORDER


def _sheet_without(ws, matrix: dict):
    ws.cell(1, 1, "Crew Without Hours").font = _TITLE_FONT
    _header_row(ws, 3, ["#", "Name", "Code", "Rank", "Base"])
    for i, c in enumerate(matrix.get("without_hours", []), start=1):
        ws.cell(3 + i, 1, i)
        ws.cell(3 + i, 2, c["name"])
        ws.cell(3 + i, 3, c["code"])
        ws.cell(3 + i, 4, c["rank"])
        ws.cell(3 + i, 5, c["base"])
    for col, w in zip("ABCDE", [6, 30, 14, 16, 10]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"


def _sheet_blocked(ws, matrix: dict):
    ws.cell(1, 1, "Blocked / Disconnected Crew").font = _TITLE_FONT
    _header_row(ws, 3, ["#", "Name", "Code", "Rank", "Reason", "Blocked On", "Status"])
    for i, c in enumerate(matrix.get("blocked", []), start=1):
        ws.cell(3 + i, 1, i)
        ws.cell(3 + i, 2, c["name"])
        ws.cell(3 + i, 3, c["code"])
        ws.cell(3 + i, 4, c["rank"])
        ws.cell(3 + i, 5, c["reason"])
        ws.cell(3 + i, 6, str(c["blocked_on"] or ""))
        ws.cell(3 + i, 7, c["status"])
    for col, w in zip("ABCDEFG", [6, 28, 14, 16, 30, 22, 14]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"


def build_statement_workbook(stmt: dict, company_name: str = "",
                             generated_by: str = "", generated_at: str = "") -> bytes:
    """Per-crew legal statement: Summary · Flight Details · Calculation Method ·
    Overrides/Audit. Every credited hour traces to a flight_id + assignment_id."""
    wb = Workbook()
    _stmt_summary(wb.active, stmt, company_name, generated_by, generated_at)
    _stmt_flights(wb.create_sheet("Flight Details"), stmt)
    _stmt_method(wb.create_sheet("Calculation Method"), stmt)
    _stmt_audit(wb.create_sheet("Overrides - Audit"), stmt)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _stmt_summary(ws, stmt, company_name, generated_by, generated_at):
    ws.title = "Summary"
    crew = stmt["crew"]
    s = stmt["summary"]
    p = stmt["period"]
    mname = calendar.month_name[p["month"]]
    ws.cell(1, 1, "Crew Hours Legal Statement").font = _TITLE_FONT
    ws.cell(2, 1, "كشف احتساب ساعات الطيران للفرد").font = _SUB_FONT

    r = 4
    for k, v in [
        ("Name", crew["name"]),
        ("Code", crew["code"]),
        ("Rank", crew["rank_code"] or crew["rank"]),
        ("Company", company_name),
        ("Base", crew["base"]),
        ("Period", f"{mname} {p['year']}"),
        ("Generated At", generated_at),
        ("Generated By", generated_by),
    ]:
        r = _kv(ws, r, k, v)

    r += 1
    ws.cell(r, 1, "Hours Summary").font = _SUB_FONT
    r += 1
    for k, v in [
        ("Operating Hours (credited)", hm(s["operating_hours"])),
        ("OFFICIAL Credited Total", hm(s["credited_total"])),
        ("Deadhead Hours (separate)", hm(s["deadhead_hours"])),
        ("Deadhead Count", s["deadhead_count"]),
        ("Standby Days", s["standby_days"]),
        ("Training Count", s["training_count"]),
        ("Observer Count", s["observer_count"]),
        ("Flights Count", s["flights_count"]),
        ("Work Days", s["work_days"]),
        ("Has Manual Overrides", "Yes" if s["has_overrides"] else "No"),
        ("DH Crediting Policy", _dh_label(s.get("dh_credit"))),
    ]:
        r = _kv(ws, r, k, v, bold_val=(k == "OFFICIAL Credited Total"))
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 36


def _stmt_flights(ws, stmt):
    headers = ["Date", "Flight No", "Duty", "From", "To", "Day Route", "Aircraft", "REG",
               "STD", "STA", "Duration", "Credited", "Included", "Reason", "Source",
               "Flight ID", "Assignment ID"]
    _header_row(ws, 1, headers)
    r = 2
    for leg in stmt["legs"]:
        vals = [leg["date"], leg["flight_no"], leg["duty_type"], leg["from"], leg["to"],
                leg.get("day_route", ""), leg["aircraft_type"], leg["registration"],
                leg["std"], leg["sta"], leg["duration_hours"], leg["credited_hours"],
                "Yes" if leg["included"] else "No", leg["reason"], leg["source"],
                leg["flight_id"], leg["assignment_id"]]
        for ci, v in enumerate(vals, start=1):
            cell = ws.cell(r, ci, v)
            if leg.get("incomplete"):
                cell.fill = PatternFill("solid", fgColor="FFF2CC")     # warning
            elif leg["included"]:
                cell.fill = _OPERATING_FILL
            else:
                cell.fill = _NOHOURS_FILL
        r += 1
    widths = [12, 10, 10, 7, 7, 16, 10, 12, 7, 7, 9, 9, 9, 34, 18, 20, 20]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(r - 1, 1)}"


def _stmt_method(ws, stmt):
    ws.cell(1, 1, "Calculation Method").font = _TITLE_FONT
    lines = [
        "Source of every credited hour: flights.duration_hours (scheduled block).",
        "Traceability chain: Monthly Total -> Day Total -> Flight Segment -> Flight Record -> Source Duration.",
        "",
        f"Crediting policy (DH crediting AT EXPORT: {_dh_label(stmt['summary'].get('dh_credit'))}):",
        "  • operating  → counted in full toward Monthly Flight Hours.",
        "  • deadhead   → credited per the DH crediting policy above (none=0% / half=50% / full=100%).",
        "  • standby    → counted as a Standby day; NOT counted as flight hours.",
        "  • training   → shown separately; not counted.",
        "  • observer   → shown separately; not counted.",
        "  • A flight with missing/zero duration is flagged and EXCLUDED until verified.",
        "",
        "Manual overrides: a super-admin may replace a day's credited hours. The original",
        "computed value, the new value, the reason, the editor, the role and the time are",
        "all retained (see the Overrides - Audit sheet). Overrides change the OFFICIAL total.",
        "",
        "Every row in 'Flight Details' carries its Flight ID and Assignment ID so the number",
        "can be reconciled against the source flight and crew-assignment records.",
    ]
    for i, line in enumerate(lines, start=3):
        ws.cell(i, 1, line)
    ws.column_dimensions["A"].width = 110


def _stmt_audit(ws, stmt):
    ws.cell(1, 1, "Manual Overrides").font = _SUB_FONT
    _header_row(ws, 2, ["Date", "Override Hours", "Old Value", "Reason", "Note", "By", "At"])
    r = 3
    for ov in stmt.get("overrides", []):
        for ci, v in enumerate([str(ov.get("duty_date") or ""), hm(ov.get("override_hours")),
                                hm(ov.get("old_value")), ov.get("reason") or "",
                                ov.get("note") or "", ov.get("created_by_name") or "",
                                str(ov.get("created_at") or "")], start=1):
            ws.cell(r, ci, v)
        r += 1

    r += 2
    ws.cell(r, 1, "Audit Log").font = _SUB_FONT
    r += 1
    _header_row(ws, r, ["At", "Action", "Old", "New", "Reason", "Note", "By", "Role"])
    r += 1
    for a in stmt.get("audit", []):
        for ci, v in enumerate([str(a.get("created_at") or ""), a.get("action") or "",
                                hm(a.get("old_value")), hm(a.get("new_value")),
                                a.get("reason") or "", a.get("note") or "",
                                a.get("performed_by_name") or "", a.get("performed_role") or ""],
                               start=1):
            ws.cell(r, ci, v)
        r += 1
    for col, w in zip("ABCDEFGH", [22, 12, 10, 10, 28, 22, 18, 14]):
        ws.column_dimensions[col].width = w


def _sheet_raw(ws, matrix: dict):
    headers = ["crew_id", "crew_name", "rank", "duty_date", "flight_no", "route",
               "aircraft_type", "registration", "std", "sta", "block_hours",
               "credited_hours", "duty_type", "notes"]
    _header_row(ws, 1, headers)
    r = 2
    y, m = matrix["year"], matrix["month"]
    for row in matrix["rows"]:
        for day_str in sorted(row["days"], key=lambda x: int(x)):
            day = row["days"][day_str]
            duty_date = f"{y:04d}-{m:02d}-{int(day_str):02d}"
            for leg in day["legs"]:
                vals = [row["crew_id"], row["name"], row["rank_code"] or row["rank"],
                        duty_date, leg["flight_no"], leg["route"], leg["aircraft_type"],
                        leg["registration"], leg["std"], leg["sta"], leg["block"],
                        leg["hours"], leg["duty_type"], ""]
                for ci, v in enumerate(vals, start=1):
                    ws.cell(r, ci, v)
                r += 1
    widths = [20, 26, 10, 12, 10, 14, 12, 14, 8, 8, 12, 14, 12, 16]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(r - 1, 1)}"
