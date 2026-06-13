"""Standby report → Excel workbook (R6.4). PURE — no DB access.

Builds a 5-sheet .xlsx from the R6.1 report dict (+ R6.2 fairness) and an
optional R6.3 roster-draft preview dict. READ-ONLY by construction: it only
formats data already computed by the read-only report path. NO financial /
payroll figures. `window_hours` is informational only (never flight hours).
"""
from __future__ import annotations
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_TITLE_FONT = Font(bold=True, size=13, color="1F4E78")
_BOLD = Font(bold=True)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _header_row(ws, row, headers):
    """Write a styled header row via explicit indices (never append)."""
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = _CENTER


def _write_row(ws, row, values):
    for col, v in enumerate(values, start=1):
        ws.cell(row=row, column=col, value=v)


def _autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _join(v) -> str:
    if isinstance(v, (list, tuple)):
        return "؛ ".join(str(x) for x in v)
    return "" if v is None else str(v)


def build_standby_workbook(report: dict, roster: dict | None = None) -> bytes:
    """report: R6.1/R6.2 dict (crew, totals, fairness, year, month, company_id).
    roster: optional R6.3 draft (slots, uncovered) — sheets stay header-only when
    None. Returns the .xlsx bytes."""
    report = report or {}
    totals = report.get("totals", {}) or {}
    fairness = report.get("fairness", {}) or {}
    wb = Workbook()

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.cell(row=1, column=1, value="تقرير الاحتياط — Standby Report").font = _TITLE_FONT
    pairs = [
        ("الشهر / Month", f"{report.get('year','')}-{str(report.get('month','')).zfill(2)}"),
        ("الشركة / Company", report.get("company_id", "")),
        ("إجمالي النوبات / Total shifts", totals.get("shifts", 0)),
        ("ساعات النوافذ (معلومة) / Window hours (info)", totals.get("window_hours", 0)),
        ("إجمالي callouts", totals.get("callouts", 0)),
        ("قبول / Accepted", totals.get("accepted", 0)),
        ("رفض / Rejected", totals.get("rejected", 0)),
        ("عدم رد / No response", totals.get("no_response", 0)),
        ("انتهاء / Expired", totals.get("expired", 0)),
        ("تعيينات ناتجة / Assignments made", totals.get("assignments_made", 0)),
        ("عدد الأفراد / Crew count", totals.get("crew_count", 0)),
    ]
    for i, (k, v) in enumerate(pairs, start=3):
        ws.cell(row=i, column=1, value=k).font = _BOLD
        ws.cell(row=i, column=2, value=v)
    _autosize(ws, [42, 28])

    # ── Sheet 2: Crew Standby Report ─────────────────────────────────────────
    ws = wb.create_sheet("Crew Standby Report")
    _header_row(ws, 1, [
        "الفرد / Crew", "الرتبة / Rank", "القاعدة / Base", "النوبات / Shifts",
        "ساعات النوافذ / Window hrs", "Callouts", "قبول / Accepted",
        "رفض / Rejected", "عدم رد / No resp", "انتهاء / Expired",
        "تعيينات / Assigned", "معدل الرد / Resp rate", "آخر callout / Last"])
    r = 2
    for c in report.get("crew", []):
        _write_row(ws, r, [
            c.get("crew_name_ar") or c.get("crew_name_en") or c.get("crew_id"),
            c.get("rank", ""), c.get("base", ""), c.get("shifts", 0),
            c.get("window_hours", 0), c.get("callouts", 0), c.get("accepted", 0),
            c.get("rejected", 0), c.get("no_response", 0), c.get("expired", 0),
            c.get("assignments_made", 0),
            "" if c.get("response_rate") is None else c.get("response_rate"),
            c.get("last_callout_at") or "",
        ])
        r += 1
    ws.freeze_panes = "A2"
    _autosize(ws, [22, 14, 10, 9, 14, 10, 10, 10, 10, 10, 10, 12, 22])

    # ── Sheet 3: Fairness ────────────────────────────────────────────────────
    ws = wb.create_sheet("Fairness")
    o = fairness.get("outliers", {}) or {}
    dist = fairness.get("distribution", {}) or {}
    av = fairness.get("averages", {}) or {}
    ws.cell(row=1, column=1, value="مقاييس العدالة / Fairness").font = _TITLE_FONT
    head = [
        ("متوسط النوبات / Avg shifts", av.get("shifts", 0)),
        ("متوسط callouts / Avg callouts", av.get("callouts", 0)),
        ("احتياط مفرط / Over standby", _join(o.get("over_standby"))),
        ("callout متكرر / Frequent callout", _join(o.get("frequent_callout"))),
        ("موثوقية منخفضة / Low reliability", _join(o.get("low_reliability"))),
        ("قواعد ناقصة التغطية / Under-covered bases", _join(o.get("under_covered_bases"))),
    ]
    for i, (k, v) in enumerate(head, start=3):
        ws.cell(row=i, column=1, value=k).font = _BOLD
        ws.cell(row=i, column=2, value=v)
    r = 3 + len(head) + 1
    for title, key in [("التوزيع حسب القاعدة / By base", "by_base"),
                       ("التوزيع حسب الرتبة / By rank", "by_rank")]:
        ws.cell(row=r, column=1, value=title).font = _BOLD
        r += 1
        _header_row(ws, r, ["", "النوبات / Shifts", "Callouts", "الأفراد / Crew"])
        r += 1
        for k, g in sorted((dist.get(key, {}) or {}).items()):
            _write_row(ws, r, [k, g.get("shifts", 0), g.get("callouts", 0),
                               g.get("crew_count", 0)])
            r += 1
        r += 1
    ws.cell(row=r, column=1, value="التوزيع حسب النوع / By type").font = _BOLD
    r += 1
    for k, v in sorted((dist.get("by_type", {}) or {}).items()):
        _write_row(ws, r, [k, v])
        r += 1
    _autosize(ws, [34, 16, 12, 12])

    # ── Sheet 4: Roster Draft ────────────────────────────────────────────────
    ws = wb.create_sheet("Roster Draft")
    _header_row(ws, 1, ["التاريخ / Date", "القاعدة / Base", "الرتبة / Rank",
                        "النوع / Type", "المرشح / Candidate", "سبب الاختيار / Reason",
                        "حِمل العدالة / Load", "تحذيرات / Warnings", "الحالة / Status"])
    r = 2
    for s in ((roster or {}).get("slots", []) or []):
        _write_row(ws, r, [
            s.get("date", ""), s.get("base", ""), s.get("rank", ""),
            s.get("standby_type", ""),
            s.get("crew_name_ar") or s.get("crew_name_en") or s.get("crew_id", ""),
            s.get("reason", ""), s.get("fairness_load", ""),
            _join(s.get("warnings")), s.get("status", "DRAFT"),
        ])
        r += 1
    ws.freeze_panes = "A2"
    _autosize(ws, [12, 10, 14, 16, 22, 26, 12, 28, 10])

    # ── Sheet 5: Uncovered Slots ─────────────────────────────────────────────
    ws = wb.create_sheet("Uncovered Slots")
    _header_row(ws, 1, ["التاريخ / Date", "القاعدة / Base", "الرتبة / Rank",
                        "النوع / Type", "سبب عدم التغطية / Reason", "تفاصيل / Details"])
    r = 2
    for u in ((roster or {}).get("uncovered", []) or []):
        _write_row(ws, r, [
            u.get("date", ""), u.get("base", ""), u.get("rank", ""),
            u.get("standby_type", ""), u.get("reason_category", ""),
            _join(u.get("reasons")),
        ])
        r += 1
    ws.freeze_panes = "A2"
    _autosize(ws, [12, 10, 14, 16, 26, 40])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
