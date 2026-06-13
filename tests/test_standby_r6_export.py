"""Reserve/Standby — R6.4 (Excel export, read-only).

POST /reports/standby/export returns an .xlsx workbook built from the SAME
R6.1/R6.2/R6.3 data path. READ-ONLY: persists nothing, no financial figures.
xlsx only (PDF deferred). Roster Draft + Uncovered sheets fill from a fresh
preview when `requirements` are supplied.

Run:  py -m pytest tests/test_standby_r6_export.py -q
"""
import asyncio
import io
from datetime import datetime, timezone

import pytest
from openpyxl import load_workbook
from fastapi import HTTPException

import app.api.v1.endpoints.standby as standby_mod
from app.core.standby_export import build_standby_workbook
from app.core.standby_report import compute_standby_report
from app.api.v1.endpoints.standby_report import standby_export
from app.core.exceptions import ForbiddenError

NOW = datetime(2026, 6, 13, 12, 0, tzinfo=timezone.utc)
SHEETS = ["Summary", "Crew Standby Report", "Fairness", "Roster Draft",
          "Uncovered Slots"]


def _wb_from_bytes(b):
    return load_workbook(io.BytesIO(b))


# ── pure workbook builder ────────────────────────────────────────────────────
def test_workbook_has_all_sheets_even_when_empty():
    wb = _wb_from_bytes(build_standby_workbook(
        compute_standby_report([], {}, NOW), None))
    assert wb.sheetnames == SHEETS


def test_workbook_numbers_match_report():
    rows = [
        {"id": "a1", "crew_id": "cr1", "standby_type": "HOME_STANDBY",
         "status": "ASSIGNED", "called_out": True, "response_status": "ACCEPTED",
         "assignment_id": "x", "called_out_at": "2026-06-05T08:30:00+03:00",
         "response_minutes": 60, "start_time": "2026-06-05T08:00:00+03:00",
         "end_time": "2026-06-05T18:00:00+03:00"},
    ]
    crew = {"cr1": {"id": "cr1", "full_name_ar": "علي", "rank": "captain", "base": "BGW"}}
    report = compute_standby_report(rows, crew, NOW)
    wb = _wb_from_bytes(build_standby_workbook(report, None))

    # Summary total shifts cell matches the report total.
    summary = wb["Summary"]
    found = {summary.cell(row=r, column=1).value: summary.cell(row=r, column=2).value
             for r in range(1, summary.max_row + 1)}
    assert found["إجمالي النوبات / Total shifts"] == report["totals"]["shifts"] == 1
    assert found["تعيينات ناتجة / Assignments made"] == 1

    # Crew sheet row matches.
    crew_ws = wb["Crew Standby Report"]
    assert crew_ws.cell(row=2, column=1).value == "علي"
    assert crew_ws.cell(row=2, column=4).value == 1            # shifts column


def test_workbook_roster_and_uncovered_sheets_fill_from_draft():
    report = compute_standby_report([], {}, NOW)
    roster = {
        "slots": [{"date": "2026-06-01", "base": "BGW", "rank": "captain",
                   "standby_type": "AIRPORT_STANDBY", "crew_name_ar": "علي",
                   "reason": "least-loaded", "fairness_load": 1,
                   "warnings": [], "status": "DRAFT"}],
        "uncovered": [{"date": "2026-06-02", "base": "BSR", "rank": "captain",
                       "standby_type": "AIRPORT_STANDBY",
                       "reason_category": "no_eligible_candidate",
                       "reasons": ["تعارض زمني"]}],
    }
    wb = _wb_from_bytes(build_standby_workbook(report, roster))
    assert wb["Roster Draft"].cell(row=2, column=9).value == "DRAFT"
    assert wb["Uncovered Slots"].cell(row=2, column=5).value == "no_eligible_candidate"
    assert "تعارض زمني" in wb["Uncovered Slots"].cell(row=2, column=6).value


# ── endpoint ─────────────────────────────────────────────────────────────────
class _R:
    def __init__(self, data): self.data = data


class _Q:
    def __init__(self, store, name):
        self.store, self.name, self._filters = store, name, []

    def select(self, *a, **k): return self
    def eq(self, f, v): self._filters.append((f, v)); return self
    def in_(self, f, vals): self._filters.append((f, list(vals))); return self
    def gte(self, *a, **k): return self
    def lt(self, *a, **k): return self
    def insert(self, p): self.store.setdefault("_writes", []).append(self.name); return self
    def update(self, p): self.store.setdefault("_writes", []).append(self.name); return self

    def _match(self, r):
        for f, v in self._filters:
            if isinstance(v, list):
                if r.get(f) not in v:
                    return False
            elif r.get(f) != v:
                return False
        return True

    def execute(self):
        return _R([dict(r) for r in self.store.get(self.name, []) if self._match(r)])


class FakeSb:
    def __init__(self, store): self.store = store
    def table(self, name): return _Q(self.store, name)


ADMIN = {"id": "u1", "role": "admin", "company_id": "c1", "is_superuser": False}
CREW = {"id": "u2", "role": "crew", "company_id": "c1", "is_superuser": False}


def _store():
    return {
        "standby_assignments": [
            {"id": "a1", "company_id": "c1", "crew_id": "cr1",
             "standby_type": "HOME_STANDBY", "status": "ASSIGNED",
             "called_out": True, "response_status": "ACCEPTED", "assignment_id": "x",
             "called_out_at": "2026-06-05T08:30:00+03:00", "response_minutes": 60,
             "start_time": "2026-06-05T08:00:00+03:00",
             "end_time": "2026-06-05T18:00:00+03:00"},
        ],
        "crew": [{"id": "cr1", "company_id": "c1", "full_name_ar": "علي",
                  "rank": "captain", "base": "BGW"}],
    }


def test_endpoint_exports_xlsx_read_only():
    store = _store()
    res = asyncio.run(standby_export(
        {"year": 2026, "month": 6}, current_user=ADMIN, sb=FakeSb(store)))
    assert res.media_type.endswith("spreadsheetml.sheet")
    assert "attachment" in res.headers["Content-Disposition"]
    wb = _wb_from_bytes(res.body)
    assert wb.sheetnames == SHEETS
    assert wb["Crew Standby Report"].cell(row=2, column=1).value == "علي"
    assert store.get("_writes") is None          # READ-ONLY


def test_endpoint_empty_month_exports_clean_file():
    store = {"standby_assignments": [], "crew": []}
    res = asyncio.run(standby_export(
        {"year": 2026, "month": 1}, current_user=ADMIN, sb=FakeSb(store)))
    wb = _wb_from_bytes(res.body)
    assert wb.sheetnames == SHEETS
    assert store.get("_writes") is None


def test_endpoint_rejects_non_xlsx_format():
    with pytest.raises(HTTPException) as ei:
        asyncio.run(standby_export(
            {"year": 2026, "month": 6, "format": "pdf"},
            current_user=ADMIN, sb=FakeSb(_store())))
    assert ei.value.status_code == 422


def test_endpoint_rbac_blocks_crew():
    with pytest.raises(ForbiddenError):
        asyncio.run(standby_export(
            {"year": 2026, "month": 6}, current_user=CREW, sb=FakeSb(_store())))


def test_endpoint_includes_roster_when_requirements_given(monkeypatch):
    monkeypatch.setattr(standby_mod, "_standby_eligibility",
                        lambda sb, c, s, e: ([], []))
    store = _store()
    res = asyncio.run(standby_export(
        {"year": 2026, "month": 6,
         "requirements": [{"base": "BGW", "rank": "captain", "per_day": 1}]},
        current_user=ADMIN, sb=FakeSb(store)))
    wb = _wb_from_bytes(res.body)
    # Roster Draft sheet now has at least one proposed slot (status DRAFT).
    assert wb["Roster Draft"].cell(row=2, column=9).value == "DRAFT"
    assert store.get("_writes") is None          # still read-only
