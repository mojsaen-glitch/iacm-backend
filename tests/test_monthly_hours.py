"""Crew Monthly Flight Hours — computation + Excel.

Run:  venv/Scripts/python -m pytest tests/test_monthly_hours.py -q
"""
import io
import openpyxl
import pytest

from datetime import datetime, timezone

from app.core.monthly_hours import (
    build_matrix, build_statement, hm, _credited_hours, _crew_type_of,
    invalidate_matrix_cache, crew_flight_hours, month_hours_by_crew,
)
from app.core.monthly_hours_excel import build_workbook, build_statement_workbook


@pytest.fixture(autouse=True)
def _clear_matrix_cache():
    # the build_matrix result cache is process-global; isolate every test.
    invalidate_matrix_cache()
    yield
    invalidate_matrix_cache()


# ── tiny Supabase stub (records the chained calls the code uses) ─────────────
class _Resp:
    def __init__(self, data):
        self.data = data


class _Query:
    def __init__(self, store, table):
        self._store, self._table = store, table
        self._in = None
        self._range = None
        self._eq = []
        self._neq = []
        self._gte = []
        self._lt = []

    @staticmethod
    def _get(row, field):
        """Resolve 'a.b' join paths into embedded dicts (PostgREST-style)."""
        cur = row
        for part in field.split("."):
            if not isinstance(cur, dict):
                return None
            cur = cur.get(part)
        return cur

    @staticmethod
    def _cmp_key(v):
        """Timestamps compare as INSTANTS (mixed +03:00 / Z offsets, like real
        Postgres); anything non-temporal falls back to string compare."""
        from datetime import datetime as _dt, timezone as _tz
        try:
            t = _dt.fromisoformat(str(v).replace("Z", "+00:00"))
            if t.tzinfo is None:
                t = t.replace(tzinfo=_tz.utc)
            return t.timestamp()
        except (ValueError, TypeError):
            return str(v)

    def select(self, *_a, **_k):
        return self

    def eq(self, field, value):
        self._eq.append((field, value))
        return self

    def neq(self, field, value):
        self._neq.append((field, value))
        return self

    def order(self, *_a, **_k):
        return self

    def gte(self, field, value):
        self._gte.append((field, value))
        return self

    def lt(self, field, value):
        self._lt.append((field, value))
        return self

    def limit(self, *_a, **_k):
        return self

    def in_(self, field, values):
        self._in = (field, set(values))
        return self

    def range(self, a, b):
        self._range = (a, b)
        return self

    def execute(self):
        data = list(self._store.get(self._table, []))
        for f, v in self._eq:
            data = [d for d in data if self._get(d, f) == v]
        for f, v in self._neq:
            data = [d for d in data if self._get(d, f) != v]
        for f, v in self._gte:
            data = [d for d in data
                    if self._cmp_key(self._get(d, f) or "") >= self._cmp_key(v)]
        for f, v in self._lt:
            data = [d for d in data
                    if self._cmp_key(self._get(d, f) or "") < self._cmp_key(v)]
        if self._in:
            f, vals = self._in
            data = [d for d in data if d.get(f) in vals]
        if self._range:
            a, b = self._range
            data = data[a:b + 1]
        return _Resp(data)


class FakeSb:
    def __init__(self, store):
        self._store = store

    def table(self, name):
        return _Query(self._store, name)


def _store():
    s = {
        "crew": [
            {"id": "c1", "full_name_en": "Capt Ali", "full_name_ar": "علي", "roster_name": "ALI",
             "employee_id": "E1", "rank": "captain", "base": "BGW", "status": "active",
             "block_reason": None, "blocked_on": None, "aircraft_qualifications": "B737",
             "max_monthly_hours": 100},
            {"id": "c2", "full_name_en": "CC Sara", "full_name_ar": "سارة", "roster_name": "SARA",
             "employee_id": "E2", "rank": "cabin_crew", "base": "BGW", "status": "active",
             "block_reason": None, "blocked_on": None, "aircraft_qualifications": "B737",
             "max_monthly_hours": 100},
            {"id": "c3", "full_name_en": "Grounded Guy", "full_name_ar": "محظور", "roster_name": "GG",
             "employee_id": "E3", "rank": "first_officer", "base": "NJF", "status": "blocked",
             "block_reason": "medical", "blocked_on": "2025-08-01T00:00:00+00:00",
             "aircraft_qualifications": "B737", "max_monthly_hours": 100},
        ],
        "flights": [
            {"id": "f1", "flight_number": "IA101", "origin_code": "BGW", "destination_code": "MED",
             "departure_time": "2025-08-03T08:00:00+00:00", "arrival_time": "2025-08-03T11:00:00+00:00",
             "duration_hours": 3.0, "aircraft_type": "B737", "aircraft_id": "a1"},
            {"id": "f2", "flight_number": "IA102", "origin_code": "BGW", "destination_code": "DXB",
             "departure_time": "2025-08-20T08:00:00+00:00", "arrival_time": "2025-08-20T12:30:00+00:00",
             "duration_hours": 4.5, "aircraft_type": "B737", "aircraft_id": "a1"},
        ],
        "aircraft": [{"id": "a1", "registration": "YI-ASA"}],
        "assignments": [
            {"id": "a1", "crew_id": "c1", "flight_id": "f1", "duty_type": "operating"},
            {"id": "a2", "crew_id": "c1", "flight_id": "f2", "duty_type": "operating"},
            {"id": "a3", "crew_id": "c2", "flight_id": "f1", "duty_type": "deadhead"},
        ],
        "companies": [{"id": "co1", "name": "Iraqi Airways"}],
    }
    for tbl in ("crew", "flights", "aircraft"):
        for row in s[tbl]:
            row.setdefault("company_id", "co1")
    return s


def test_hm():
    assert hm(7.5) == "7:30"
    assert hm(4.5) == "4:30"
    assert hm(0) == ""
    assert hm(1.25) == "1:15"


def test_credited_policy():
    assert _credited_hours("operating", 3.0) == 3.0
    assert _credited_hours("deadhead", 3.0) == 0.0
    assert _credited_hours("standby", 3.0) == 0.0


def test_crew_type():
    assert _crew_type_of("captain") == "pilots"
    assert _crew_type_of("cabin_crew") == "cabin"


def test_build_matrix_aggregates():
    m = build_matrix(FakeSb(_store()), "co1", 2025, 8, {})
    assert m["days_in_month"] == 31
    rows = {r["crew_id"]: r for r in m["rows"]}

    cap = rows["c1"]
    assert cap["month_total"] == 7.5
    assert cap["first_half"] == 3.0      # day 3
    assert cap["second_half"] == 4.5     # day 20
    assert cap["flights_count"] == 2
    assert cap["work_days"] == 2
    assert "3" in cap["days"] and "20" in cap["days"]
    assert cap["days"]["3"]["legs"][0]["route"] == "BGW-MED"
    assert cap["days"]["3"]["legs"][0]["registration"] == "YI-ASA"

    cc = rows["c2"]
    assert cc["month_total"] == 0.0          # deadhead not credited
    assert cc["deadhead_count"] == 1

    gg = rows["c3"]
    assert gg["blocked"] is True

    s = m["summary"]
    assert s["active_crew"] == 1             # only the captain has hours
    assert s["total_flights"] == 2
    assert s["blocked_crew"] == 1
    assert s["total_hours"] == 7.5

    assert any(b["crew_id"] == "c3" for b in m["blocked"])


def test_cancelled_flight_not_credited():
    """A CANCELLED flight must neither credit hours nor appear as a duty cell —
    engine-level exclusion (matrix + Excel + crew profile all agree)."""
    store = _store()
    # 2h operating leg (counts) + 5h CANCELLED leg (must NOT count) for c1.
    store["flights"] += [
        {"id": "f_ok", "flight_number": "IA103", "origin_code": "BGW",
         "destination_code": "EBL", "company_id": "co1",
         "departure_time": "2025-08-10T08:00:00+00:00",
         "arrival_time": "2025-08-10T10:00:00+00:00",
         "duration_hours": 2.0, "aircraft_type": "B737", "aircraft_id": "a1"},
        {"id": "f_cx", "flight_number": "IA104", "origin_code": "BGW",
         "destination_code": "AMM", "company_id": "co1", "status": "cancelled",
         "departure_time": "2025-08-11T08:00:00+00:00",
         "arrival_time": "2025-08-11T13:00:00+00:00",
         "duration_hours": 5.0, "aircraft_type": "B737", "aircraft_id": "a1"},
    ]
    store["assignments"] += [
        {"id": "a_ok", "crew_id": "c1", "flight_id": "f_ok", "duty_type": "operating"},
        {"id": "a_cx", "crew_id": "c1", "flight_id": "f_cx", "duty_type": "operating"},
    ]
    invalidate_matrix_cache()
    m = build_matrix(FakeSb(store), "co1", 2025, 8, {})
    cap = {r["crew_id"]: r for r in m["rows"]}["c1"]
    assert cap["month_total"] == 9.5          # 3.0 + 4.5 + 2.0 — NOT the 5h cancelled
    assert "11" not in cap["days"]            # cancelled leg leaves no duty cell
    # Deadhead policy unchanged: c2 still credits nothing.
    assert {r["crew_id"]: r for r in m["rows"]}["c2"]["month_total"] == 0.0


def test_crew_profile_excludes_cancelled_too():
    """GET /crew/{id}/flight-hours path (crew_flight_hours) — same engine rule."""
    store = _store()
    store["flights"].append(
        {"id": "f_cx", "flight_number": "IA104", "origin_code": "BGW",
         "destination_code": "AMM", "company_id": "co1", "status": "cancelled",
         "departure_time": "2025-08-11T08:00:00+00:00",
         "arrival_time": "2025-08-11T13:00:00+00:00",
         "duration_hours": 5.0, "aircraft_type": "B737", "aircraft_id": "a1"})
    store["assignments"].append(
        {"id": "a_cx", "crew_id": "c1", "flight_id": "f_cx", "duty_type": "operating"})
    res = crew_flight_hours(FakeSb(store), "co1", "c1")
    assert res["total"] == 7.5                # 3.0 + 4.5 — cancelled 5h excluded
    # Deadhead-only crew stays at zero (policy untouched).
    assert crew_flight_hours(FakeSb(store), "co1", "c2")["total"] == 0.0


# ── Baghdad month bounds (official reports) ───────────────────────────────────
def _boundary_store():
    """Crew fixture + a red-eye departing 2026-06-30T22:00Z = 01:00 JULY 1
    Baghdad — must belong to JULY, day cell '1'."""
    s = _store()
    s["flights"] = [
        # 01:00 Jul 1 Baghdad (22:00Z Jun 30) → JULY, day 1
        {"id": "fb", "flight_number": "IA201", "origin_code": "BGW",
         "destination_code": "IST", "company_id": "co1",
         "departure_time": "2026-06-30T22:00:00+00:00",
         "arrival_time": "2026-07-01T01:00:00+00:00",
         "duration_hours": 2.0, "aircraft_type": "B737", "aircraft_id": "a1"},
        # same boundary instant but CANCELLED → never credited
        {"id": "fbx", "flight_number": "IA202", "origin_code": "BGW",
         "destination_code": "AMM", "company_id": "co1", "status": "cancelled",
         "departure_time": "2026-06-30T22:30:00+00:00",
         "arrival_time": "2026-07-01T02:30:00+00:00",
         "duration_hours": 5.0, "aircraft_type": "B737", "aircraft_id": "a1"},
        # 21:00 Jul 31 Baghdad (18:00Z) → stays in JULY, day 31
        {"id": "fl", "flight_number": "IA203", "origin_code": "BGW",
         "destination_code": "EBL", "company_id": "co1",
         "departure_time": "2026-07-31T18:00:00+00:00",
         "arrival_time": "2026-07-31T19:30:00+00:00",
         "duration_hours": 1.5, "aircraft_type": "B737", "aircraft_id": "a1"},
    ]
    s["assignments"] = [
        {"id": "ab", "crew_id": "c1", "flight_id": "fb", "duty_type": "operating"},
        {"id": "abx", "crew_id": "c1", "flight_id": "fbx", "duty_type": "operating"},
        {"id": "al", "crew_id": "c1", "flight_id": "fl", "duty_type": "operating"},
        {"id": "ad", "crew_id": "c2", "flight_id": "fb", "duty_type": "deadhead"},
    ]
    return s


def test_baghdad_boundary_flight_counts_in_july_not_june():
    store = _boundary_store()
    invalidate_matrix_cache()
    july = build_matrix(FakeSb(store), "co1", 2026, 7, {})
    cap = {r["crew_id"]: r for r in july["rows"]}["c1"]
    assert cap["month_total"] == 3.5            # 2.0 (red-eye) + 1.5 (last day)
    assert "1" in cap["days"]                   # Baghdad day cell, NOT June 30
    assert "31" in cap["days"]                  # last-day flight stays in July
    # Cancelled boundary leg neither credits nor appears.
    assert all(l["flight_no"] != "IA202"
               for d in cap["days"].values() for l in d["legs"])
    # Deadhead on the boundary still credits nothing.
    assert {r["crew_id"]: r for r in july["rows"]}["c2"]["month_total"] == 0.0

    june = build_matrix(FakeSb(store), "co1", 2026, 6, {})
    june_rows = {r["crew_id"]: r for r in june["rows"]}
    assert june_rows.get("c1") is None or june_rows["c1"]["month_total"] == 0.0


def test_baghdad_profile_matches_matrix():
    """crew_flight_hours (the profile) buckets the CURRENT Baghdad month the
    same way the matrix does — boundary flight included in both."""
    from datetime import timedelta
    now = datetime.now(timezone.utc).replace(tzinfo=None)
    bag_now = now + timedelta(hours=3)
    month_start_utc = bag_now.replace(day=1, hour=0, minute=0,
                                      second=0, microsecond=0) - timedelta(hours=3)
    dep_in = (month_start_utc + timedelta(hours=1)).strftime("%Y-%m-%dT%H:%M:%S+00:00")
    dep_out = (month_start_utc - timedelta(hours=1)).strftime("%Y-%m-%dT%H:%M:%S+00:00")
    store = _store()
    store["flights"] = [
        {"id": "fin", "flight_number": "IA301", "origin_code": "BGW",
         "destination_code": "IST", "company_id": "co1",
         "departure_time": dep_in, "arrival_time": dep_in,
         "duration_hours": 2.0, "aircraft_type": "B737", "aircraft_id": "a1"},
        {"id": "fout", "flight_number": "IA302", "origin_code": "BGW",
         "destination_code": "EBL", "company_id": "co1",
         "departure_time": dep_out, "arrival_time": dep_out,
         "duration_hours": 7.0, "aircraft_type": "B737", "aircraft_id": "a1"},
    ]
    store["assignments"] = [
        {"id": "ain", "crew_id": "c1", "flight_id": "fin", "duty_type": "operating"},
        {"id": "aout", "crew_id": "c1", "flight_id": "fout", "duty_type": "operating"},
    ]
    profile = crew_flight_hours(FakeSb(store), "co1", "c1")
    assert profile["month"] == 2.0              # only the in-month boundary leg
    invalidate_matrix_cache()
    m = build_matrix(FakeSb(store), "co1", bag_now.year, bag_now.month, {})
    cap = {r["crew_id"]: r for r in m["rows"]}["c1"]
    assert cap["month_total"] == profile["month"]   # matrix == profile


def test_month_hours_by_crew_baghdad_boundary():
    from datetime import timedelta
    now = datetime.now(timezone.utc).replace(tzinfo=None)
    bag_now = now + timedelta(hours=3)
    month_start_utc = bag_now.replace(day=1, hour=0, minute=0,
                                      second=0, microsecond=0) - timedelta(hours=3)
    dep_in = (month_start_utc + timedelta(hours=1)).strftime("%Y-%m-%dT%H:%M:%S+00:00")
    dep_out = (month_start_utc - timedelta(hours=1)).strftime("%Y-%m-%dT%H:%M:%S+00:00")
    store = {"assignments": [
        {"crew_id": "c1", "duty_type": "operating",
         "flights": {"duration_hours": 2.0, "departure_time": dep_in,
                     "status": "scheduled", "company_id": "co1"}},
        {"crew_id": "c1", "duty_type": "operating",
         "flights": {"duration_hours": 7.0, "departure_time": dep_out,
                     "status": "scheduled", "company_id": "co1"}},
    ]}
    out = month_hours_by_crew(FakeSb(store), "co1")
    assert out.get("c1") == 2.0                 # previous-Baghdad-month leg excluded


def test_month_hours_by_crew_batch():
    """The IROPS-fairness batch: real credited hours per crew in one join —
    same policy (operating full, deadhead 0, cancelled excluded)."""
    now = datetime.now(timezone.utc)
    this_month = now.replace(day=2).strftime("%Y-%m-%dT08:00:00+00:00")
    store = {
        "assignments": [
            # operating 3h this month → credits
            {"crew_id": "c1", "duty_type": "operating",
             "flights": {"duration_hours": 3.0, "departure_time": this_month,
                         "status": "scheduled", "company_id": "co1"}},
            # CANCELLED 5h → must NOT credit
            {"crew_id": "c1", "duty_type": "operating",
             "flights": {"duration_hours": 5.0, "departure_time": this_month,
                         "status": "cancelled", "company_id": "co1"}},
            # deadhead 4h → policy unchanged: 0
            {"crew_id": "c2", "duty_type": "deadhead",
             "flights": {"duration_hours": 4.0, "departure_time": this_month,
                         "status": "scheduled", "company_id": "co1"}},
        ],
    }
    out = month_hours_by_crew(FakeSb(store), "co1")
    assert out.get("c1") == 3.0          # not 8.0 — cancelled excluded
    assert out.get("c2", 0.0) == 0.0     # deadhead still credits nothing


def test_only_with_hours_filter():
    m = build_matrix(FakeSb(_store()), "co1", 2025, 8, {"only_with_hours": True})
    ids = {r["crew_id"] for r in m["rows"]}
    assert ids == {"c1"}                      # only the captain


def test_override_applies():
    store = _store()
    store["crew_hours_overrides"] = [
        {"company_id": "co1", "crew_id": "c1", "duty_date": "2025-08-03", "override_hours": 9.0},
    ]
    m = build_matrix(FakeSb(store), "co1", 2025, 8, {})
    cap = {r["crew_id"]: r for r in m["rows"]}["c1"]
    assert cap["days"]["3"]["day_hours"] == 9.0
    assert cap["days"]["3"]["override"] is True
    assert cap["days"]["3"]["computed_hours"] == 3.0     # original preserved
    assert cap["first_half"] == 9.0                      # totals use the override
    assert cap["month_total"] == 13.5                    # 9.0 + 4.5
    assert cap["has_overrides"] is True


def test_build_statement_traceable():
    m = build_statement(FakeSb(_store()), "co1", "c1", 2025, 8)
    assert m["crew"]["crew_id"] == "c1"
    assert m["summary"]["operating_hours"] == 7.5
    assert m["summary"]["credited_total"] == 7.5
    assert m["summary"]["flights_count"] == 2
    op = [leg for leg in m["legs"] if leg["duty_type"] == "operating"]
    assert len(op) == 2
    # every credited leg traces to a flight + assignment + source
    assert all(leg["included"] and leg["flight_id"] and leg["assignment_id"] for leg in op)
    assert all(leg["source"] == "flights.duration_hours" for leg in m["legs"])
    # statement workbook has the 4 legal sheets
    data = build_statement_workbook(m, "Iraqi Airways", "Admin", "2026-06-04T00:00:00+00:00")
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["Summary", "Flight Details", "Calculation Method", "Overrides - Audit"]


def test_statement_reflects_override():
    store = _store()
    store["crew_hours_overrides"] = [
        {"company_id": "co1", "crew_id": "c1", "duty_date": "2025-08-03", "override_hours": 9.0,
         "old_value": 3.0, "reason": "manual", "note": "", "created_by_name": "Admin", "created_at": "x"},
    ]
    m = build_statement(FakeSb(store), "co1", "c1", 2025, 8)
    assert m["summary"]["operating_hours"] == 7.5     # computed stays
    assert m["summary"]["credited_total"] == 13.5     # official reflects override (9 + 4.5)
    assert m["summary"]["has_overrides"] is True
    assert len(m["overrides"]) == 1


def test_dh_credit_full():
    # with full DH crediting the deadhead crew (c2) now earns credited hours,
    # while operating crew are unchanged and DH is still reported separately.
    m = build_matrix(FakeSb(_store()), "co1", 2025, 8, {"dh_credit": "full"})
    rows = {r["crew_id"]: r for r in m["rows"]}
    assert rows["c2"]["month_total"] == 3.0
    assert rows["c2"]["deadhead_count"] == 1
    assert rows["c1"]["month_total"] == 7.5
    assert m["summary"]["dh_credit"] == "full"


def test_summary_breakdowns():
    s = build_matrix(FakeSb(_store()), "co1", 2025, 8, {})["summary"]
    assert s["top10"][0]["hours"] == 7.5
    assert any(b["key"] for b in s["by_rank"])
    assert isinstance(s["by_aircraft"], list)
    assert s["dh_credit"] == "none"


def test_statement_dh_credit():
    m = build_statement(FakeSb(_store()), "co1", "c2", 2025, 8, dh_credit="full")
    assert m["summary"]["operating_hours"] == 0.0      # c2 has no operating legs
    assert m["summary"]["credited_total"] == 3.0       # deadhead credited at 100%
    dh = [leg for leg in m["legs"] if leg["duty_type"] == "deadhead"]
    assert dh and dh[0]["included"] is True and dh[0]["credited_hours"] == 3.0


def test_build_workbook_valid_xlsx():
    m = build_matrix(FakeSb(_store()), "co1", 2025, 8, {})
    data = build_workbook(m, "Iraqi Airways")
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert wb.sheetnames == [
        "Monthly Matrix", "Summary", "Calculation Details",
        "Crew Without Hours", "Blocked - Disconnected", "Raw Data",
    ]
    # Raw Data should have at least the captain's 2 operating legs + header.
    raw = wb["Raw Data"]
    assert raw.max_row >= 3


def _all_text(ws):
    return "\n".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)


def test_matrix_cache_and_invalidation():
    import app.core.monthly_hours as mh
    store = _store()
    sb = FakeSb(store)
    m1 = mh.build_matrix(sb, "co1", 2025, 8, {})
    # mutate the source AFTER first call → cached result is returned unchanged
    store["crew_hours_overrides"] = [{"company_id": "co1", "crew_id": "c1",
                                      "duty_date": "2025-08-03", "override_hours": 9.0}]
    m2 = mh.build_matrix(sb, "co1", 2025, 8, {})
    assert m2 is m1                                   # served from cache
    mh.invalidate_matrix_cache("co1")
    m3 = mh.build_matrix(sb, "co1", 2025, 8, {})
    cap = {r["crew_id"]: r for r in m3["rows"]}["c1"]
    assert cap["month_total"] == 13.5                # recomputed with the override


def test_dh_policy_stamped_in_exports():
    # The crediting policy (DH credit) must be visible in every exported report.
    m = build_matrix(FakeSb(_store()), "co1", 2025, 8, {"dh_credit": "full"})
    wb = openpyxl.load_workbook(io.BytesIO(build_workbook(m, "Iraqi Airways")))
    summary_text = _all_text(wb["Summary"])
    assert "DH Crediting Policy" in summary_text
    assert "Full (100%)" in summary_text
    matrix_head = "\n".join(
        str(c.value) for row in wb["Monthly Matrix"].iter_rows(max_row=3) for c in row if c.value)
    assert "DH Crediting" in matrix_head

    st = build_statement(FakeSb(_store()), "co1", "c2", 2025, 8, dh_credit="full")
    swb = openpyxl.load_workbook(io.BytesIO(build_statement_workbook(st, "Iraqi Airways", "Admin", "x")))
    assert "Full (100%)" in _all_text(swb["Calculation Method"])
    assert "Full (100%)" in _all_text(swb["Summary"])
